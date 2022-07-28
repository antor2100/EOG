#from openpyxl import load_workbook
import openpyxl
import csv
import sys

print(sys.argv[2])

if sys.argv[2] == "1":
	name = "snpp_mooncycles.xlsx"
elif sys.argv[2] == "2":
	name = "j01_mooncycles.xlsx"
print(name)
wb = openpyxl.load_workbook(filename = name)

try:
	del wb[sys.argv[1]]
except:
	print("sheet not in list")

ws = wb.active
ws = wb.create_sheet(title=sys.argv[1])

with open('granule_summary.csv') as file_obj:
    reader_obj = csv.reader(file_obj)
      
    for row in reader_obj:
        ws.append(row)

img = openpyxl.drawing.image.Image('images/smi.png')
img.width = 700
img.height = 512
img.anchor = 'B10'
ws.add_image(img)

img = openpyxl.drawing.image.Image('images/std.png')
img.width = 700
img.height = 512
img.anchor = 'B37'
ws.add_image(img)

img = openpyxl.drawing.image.Image('images/bowtie.png')
img.width = 700
img.height = 512
img.anchor = 'B64'
ws.add_image(img)

img = openpyxl.drawing.image.Image('images/concat_smi.png')
img.width = 700
img.height = 512
img.anchor = 'M10'
ws.add_image(img)

img = openpyxl.drawing.image.Image('images/concat_std.png')
img.width = 700
img.height = 512
img.anchor = 'M37'
ws.add_image(img)

img = openpyxl.drawing.image.Image('images/concat_bowtie.png')
img.width = 700
img.height = 512
img.anchor = 'M64'
ws.add_image(img)

wb.save(filename = name)
